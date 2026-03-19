"""
Reverse DCF Screener
====================
Batch-run the Reverse DCF engine across multiple tickers.
Produces a ranked screening table with implied growth, upside, and flags.

Usage:
    python screener.py data_folder/ --sort upside
    python screener.py data_folder/ --export screening_results.xlsx
"""

import argparse
from pathlib import Path
import pandas as pd
import numpy as np
from reverse_dcf_engine import ReverseDCF, DCFParams


def screen_folder(folder: str, params: DCFParams = None) -> pd.DataFrame:
    """Run Reverse DCF on all Excel files in a folder."""
    folder = Path(folder)
    files = sorted(folder.glob("rdcf_data_*.xlsx"))

    if not files:
        print(f"No rdcf_data_*.xlsx files found in {folder}")
        return pd.DataFrame()

    results = []
    for f in files:
        try:
            model = ReverseDCF.from_excel(str(f), params)
            r = model.run()

            base = r["scenarios"]["Base"]
            bull = r["scenarios"]["Bull"]
            bear = r["scenarios"]["Bear"]

            results.append({
                "Ticker": r["ticker"],
                "Price": r["price"],
                "Implied_Growth": r["implied_growth"],
                "WACC": r["wacc"],
                "EBIT_Margin": r["ebit_margin_used"],
                "ROIC": r["roic_gate"]["roic"],
                "ROIC_WACC_Spread": r["roic_gate"]["spread"],
                "Value_Creating": r["roic_gate"]["value_creating"],
                "TV_Pct": r["tv_decomposition"]["tv_pct"],
                "Bull_Fair_Price": bull["fair_price"],
                "Base_Fair_Price": base["fair_price"],
                "Bear_Fair_Price": bear["fair_price"],
                "Bull_Upside": bull["upside_downside"],
                "Base_Upside": base["upside_downside"],
                "Bear_Upside": bear["upside_downside"],
                "Hist_5Y_CAGR": r["historical_profile"].revenue_cagr_5y,
                "Implied_vs_5Y": (
                    r["implied_growth"] / r["historical_profile"].revenue_cagr_5y
                    if r["historical_profile"].revenue_cagr_5y != 0 else np.nan
                ),
                "Plausibility_Flags": sum(
                    1 for c in r["plausibility"] if c["flag"] == "🔴"
                ),
            })
            print(f"  ✓ {r['ticker']}")
        except Exception as e:
            print(f"  ✗ {f.name}: {e}")

    df = pd.DataFrame(results)
    return df


def rank_screen(df: pd.DataFrame, sort_by: str = "Base_Upside") -> pd.DataFrame:
    """Rank screening results."""
    if df.empty:
        return df

    df = df.sort_values(sort_by, ascending=False).reset_index(drop=True)
    df.index = df.index + 1
    df.index.name = "Rank"

    # Composite signal: positive upside + value-creating + plausible
    df["Signal"] = np.where(
        (df["Base_Upside"] > 0) & (df["Value_Creating"]) & (df["Plausibility_Flags"] == 0),
        "🟢 BUY",
        np.where(
            (df["Base_Upside"] > 0) & (df["Plausibility_Flags"] <= 1),
            "🟡 WATCH",
            np.where(
                df["Base_Upside"] < -0.10,
                "🔴 EXPENSIVE",
                "⚪ NEUTRAL"
            )
        )
    )
    return df


def export_screening(df: pd.DataFrame, output: str):
    """Export to formatted Excel."""
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Screening", index=True)
        ws = writer.sheets["Screening"]

        # Basic formatting
        from openpyxl.styles import Font, PatternFill, Alignment, numbers

        header_fill = PatternFill("solid", fgColor="003850")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Format percentage columns
        pct_cols = [c for c in df.columns if any(
            k in c for k in ["Growth", "Margin", "ROIC", "Spread", "Upside", "WACC", "CAGR", "TV_Pct"]
        )]
        for col_name in pct_cols:
            col_idx = list(df.columns).index(col_name) + 2  # +1 for rank, +1 for 1-based
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = "0.0%"

        # Price columns
        price_cols = [c for c in df.columns if "Price" in c]
        for col_name in price_cols:
            col_idx = list(df.columns).index(col_name) + 2
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = "#,##0.00"

        # Column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 18)

    print(f"Exported: {output}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Reverse DCF Screener")
    parser.add_argument("folder", help="Folder with rdcf_data_*.xlsx files")
    parser.add_argument("--sort", default="Base_Upside", help="Sort column (default: Base_Upside)")
    parser.add_argument("--export", default="", help="Export to Excel path")
    parser.add_argument("--wacc", type=float, default=None, help="Override WACC for all tickers")
    args = parser.parse_args()

    params = DCFParams()
    if args.wacc:
        params.wacc_override = args.wacc

    print(f"Screening {args.folder}...\n")
    df = screen_folder(args.folder, params)
    df = rank_screen(df, args.sort)

    if not df.empty:
        # Print summary
        fmt_cols = ["Ticker", "Price", "Implied_Growth", "ROIC", "Base_Upside", "TV_Pct", "Signal"]
        print(f"\n{df[fmt_cols].to_string()}\n")

        if args.export:
            export_screening(df, args.export)
